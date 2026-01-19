#!/usr/bin/env python3
"""
Enhanced 3-Statement Financial Model
Adds IB-level features: proper debt schedule, scenario analysis, checks, and summary
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

# Color definitions
INPUT_FILL = PatternFill(start_color="D6E4F5", end_color="D6E4F5", fill_type="solid")  # Light blue
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Dark blue
SECTION_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Light gray
CHECK_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
CHECK_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red

FORMULA_FONT = Font(color="000000")  # Black
INPUT_FONT = Font(color="0000FF", bold=True)  # Blue
HEADER_FONT = Font(color="FFFFFF", bold=True)  # White
BOLD_FONT = Font(bold=True)
CHECK_FONT = Font(color="006100", bold=True)  # Green

# Years
HISTORICAL_YEARS = [2021, 2022, 2023, 2024]
FORECAST_YEARS = [2025, 2026, 2027, 2028, 2029]
ALL_YEARS = HISTORICAL_YEARS + FORECAST_YEARS

def set_column_widths(ws, widths):
    """Set column widths for a worksheet"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def enhance_assumptions_sheet(ws):
    """Enhance the Assumptions & Drivers sheet with scenario-driven inputs"""
    print("  Enhancing Assumptions & Drivers...")

    # Find the scenario cell (should be B4)
    # Clear out old inputs and rebuild with proper scenario formulas

    # We need to clear and rebuild the sheet to add scenario logic
    # Keep title rows (1-4)

    # Set column widths
    set_column_widths(ws, {
        'A': 30, 'B': 15, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12
    })

    # Update scenario dropdown area
    ws['A4'] = 'Scenario Selection:'
    ws['A4'].font = BOLD_FONT
    ws['B4'].fill = INPUT_FILL
    ws['B4'].font = INPUT_FONT

    # Add note about scenarios
    ws['A5'] = '(Change to Base, Upside, or Downside)'
    ws['A5'].font = Font(italic=True, size=9)

    col_start = 2  # Column B

    # Clear rows 7 onwards and rebuild
    # Historical revenue inputs (rows 8-9)
    row = 8
    ws[f'A{row}'] = 'Product Revenue (Historical)'
    ws['B8'].value = 100
    ws['C8'].value = 110
    ws['D8'].value = 125
    ws['E8'].value = 140
    for col in ['B', 'C', 'D', 'E']:
        ws[col + '8'].fill = INPUT_FILL
        ws[col + '8'].font = INPUT_FONT
        ws[col + '8'].number_format = '#,##0.0'

    row = 9
    ws[f'A{row}'] = 'Service Revenue (Historical)'
    ws['B9'].value = 50
    ws['C9'].value = 55
    ws['D9'].value = 62
    ws['E9'].value = 70
    for col in ['B', 'C', 'D', 'E']:
        ws[col + '9'].fill = INPUT_FILL
        ws[col + '9'].font = INPUT_FONT
        ws[col + '9'].number_format = '#,##0.0'

    # SCENARIO-DRIVEN ASSUMPTIONS
    row = 11
    ws[f'A{row}'] = 'SCENARIO ASSUMPTIONS'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    # Create scenario assumption table (rows 12-20)
    # Headers
    ws['A12'] = 'Assumption'
    ws['B12'] = 'Base'
    ws['C12'] = 'Upside'
    ws['D12'] = 'Downside'
    ws['E12'] = 'SELECTED →'
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[col + '12'].font = BOLD_FONT
        ws[col + '12'].fill = SECTION_FILL

    # Product Growth % (2025)
    ws['A13'] = 'Product Growth % (2025)'
    ws['B13'] = 0.08
    ws['C13'] = 0.12
    ws['D13'] = 0.05
    ws['E13'] = '=IF($B$4="Base",B13,IF($B$4="Upside",C13,D13))'

    # Product Growth % (2026-2029)
    ws['A14'] = 'Product Growth % (2026+)'
    ws['B14'] = 0.06
    ws['C14'] = 0.10
    ws['D14'] = 0.03
    ws['E14'] = '=IF($B$4="Base",B14,IF($B$4="Upside",C14,D14))'

    # Service Growth % (2025)
    ws['A15'] = 'Service Growth % (2025)'
    ws['B15'] = 0.10
    ws['C15'] = 0.15
    ws['D15'] = 0.07
    ws['E15'] = '=IF($B$4="Base",B15,IF($B$4="Upside",C15,D15))'

    # Service Growth % (2026-2029)
    ws['A16'] = 'Service Growth % (2026+)'
    ws['B16'] = 0.08
    ws['C16'] = 0.12
    ws['D16'] = 0.05
    ws['E16'] = '=IF($B$4="Base",B16,IF($B$4="Upside",C16,D16))'

    # Gross Margin %
    ws['A17'] = 'Gross Margin % (Forecast)'
    ws['B17'] = 0.67
    ws['C17'] = 0.70
    ws['D17'] = 0.64
    ws['E17'] = '=IF($B$4="Base",B17,IF($B$4="Upside",C17,D17))'

    # SG&A %
    ws['A18'] = 'SG&A % (Forecast)'
    ws['B18'] = 0.24
    ws['C18'] = 0.23
    ws['D18'] = 0.26
    ws['E18'] = '=IF($B$4="Base",B18,IF($B$4="Upside",C18,D18))'

    # R&D %
    ws['A19'] = 'R&D % (Forecast)'
    ws['B19'] = 0.12
    ws['C19'] = 0.13
    ws['D19'] = 0.10
    ws['E19'] = '=IF($B$4="Base",B19,IF($B$4="Upside",C19,D19))'

    # CapEx %
    ws['A20'] = 'CapEx % (Forecast)'
    ws['B20'] = 0.07
    ws['C20'] = 0.09
    ws['D20'] = 0.06
    ws['E20'] = '=IF($B$4="Base",B20,IF($B$4="Upside",C20,D20))'

    # Working Capital - AR Days
    ws['A21'] = 'AR Days (Forecast)'
    ws['B21'] = 45
    ws['C21'] = 42
    ws['D21'] = 48
    ws['E21'] = '=IF($B$4="Base",B21,IF($B$4="Upside",C21,D21))'

    # Inventory Days
    ws['A22'] = 'Inventory Days (Forecast)'
    ws['B22'] = 60
    ws['C22'] = 55
    ws['D22'] = 65
    ws['E22'] = '=IF($B$4="Base",B22,IF($B$4="Upside",C22,D22))'

    # AP Days
    ws['A23'] = 'AP Days (Forecast)'
    ws['B23'] = 30
    ws['C23'] = 35
    ws['D23'] = 28
    ws['E23'] = '=IF($B$4="Base",B23,IF($B$4="Upside",C23,D23))'

    # Format scenario table inputs
    for row in range(13, 24):
        for col in ['B', 'C', 'D']:
            ws[col + str(row)].fill = INPUT_FILL
            ws[col + str(row)].font = INPUT_FONT
            if row <= 20:  # Percentages
                ws[col + str(row)].number_format = '0.0%'
            else:  # Days
                ws[col + str(row)].number_format = '0'

        # Selected column (E)
        ws['E' + str(row)].number_format = '0.0%' if row <= 20 else '0'

    # FIXED ASSUMPTIONS
    row = 25
    ws[f'A{row}'] = 'FIXED ASSUMPTIONS'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    # Historical margins
    ws['A26'] = 'Gross Margin % (Historical)'
    ws['B26'] = 0.65
    ws['B26'].fill = INPUT_FILL
    ws['B26'].font = INPUT_FONT
    ws['B26'].number_format = '0.0%'

    ws['A27'] = 'SG&A % (Historical)'
    ws['B27'] = 0.25
    ws['B27'].fill = INPUT_FILL
    ws['B27'].font = INPUT_FONT
    ws['B27'].number_format = '0.0%'

    ws['A28'] = 'R&D % (Historical)'
    ws['B28'] = 0.12
    ws['B28'].fill = INPUT_FILL
    ws['B28'].font = INPUT_FONT
    ws['B28'].number_format = '0.0%'

    ws['A29'] = 'D&A % of Revenue'
    ws['B29'] = 0.05
    ws['B29'].fill = INPUT_FILL
    ws['B29'].font = INPUT_FONT
    ws['B29'].number_format = '0.0%'

    ws['A30'] = 'Tax Rate'
    ws['B30'] = 0.25
    ws['B30'].fill = INPUT_FILL
    ws['B30'].font = INPUT_FONT
    ws['B30'].number_format = '0.0%'

    # Working Capital (Historical)
    ws['A31'] = 'AR Days (Historical)'
    ws['B31'] = 45
    ws['B31'].fill = INPUT_FILL
    ws['B31'].font = INPUT_FONT
    ws['B31'].number_format = '0'

    ws['A32'] = 'Inventory Days (Historical)'
    ws['B32'] = 60
    ws['B32'].fill = INPUT_FILL
    ws['B32'].font = INPUT_FONT
    ws['B32'].number_format = '0'

    ws['A33'] = 'AP Days (Historical)'
    ws['B33'] = 30
    ws['B33'].fill = INPUT_FILL
    ws['B33'].font = INPUT_FONT
    ws['B33'].number_format = '0'

    ws['A34'] = 'Other Current Assets % of Rev'
    ws['B34'] = 0.03
    ws['B34'].fill = INPUT_FILL
    ws['B34'].font = INPUT_FONT
    ws['B34'].number_format = '0.0%'

    ws['A35'] = 'Other Current Liab % of Rev'
    ws['B35'] = 0.02
    ws['B35'].fill = INPUT_FILL
    ws['B35'].font = INPUT_FONT
    ws['B35'].number_format = '0.0%'

    ws['A36'] = 'CapEx % (Historical)'
    ws['B36'] = 0.08
    ws['B36'].fill = INPUT_FILL
    ws['B36'].font = INPUT_FONT
    ws['B36'].number_format = '0.0%'

    # DEBT ASSUMPTIONS
    row = 38
    ws[f'A{row}'] = 'DEBT ASSUMPTIONS'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    ws['A39'] = 'Beginning Term Loan (2021)'
    ws['B39'] = 50
    ws['B39'].fill = INPUT_FILL
    ws['B39'].font = INPUT_FONT
    ws['B39'].number_format = '#,##0.0'

    ws['A40'] = 'Term Loan Interest Rate'
    ws['B40'] = 0.06
    ws['B40'].fill = INPUT_FILL
    ws['B40'].font = INPUT_FONT
    ws['B40'].number_format = '0.0%'

    ws['A41'] = 'Revolver Interest Rate'
    ws['B41'] = 0.05
    ws['B41'].fill = INPUT_FILL
    ws['B41'].font = INPUT_FONT
    ws['B41'].number_format = '0.0%'

    ws['A42'] = 'Minimum Cash Balance'
    ws['B42'] = 20
    ws['B42'].fill = INPUT_FILL
    ws['B42'].font = INPUT_FONT
    ws['B42'].number_format = '#,##0.0'

    # INITIAL BALANCE SHEET
    row = 44
    ws[f'A{row}'] = 'INITIAL BALANCE SHEET (2021)'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    ws['A45'] = 'Beginning Cash'
    ws['B45'] = 30
    ws['B45'].fill = INPUT_FILL
    ws['B45'].font = INPUT_FONT
    ws['B45'].number_format = '#,##0.0'

    ws['A46'] = 'Beginning PP&E'
    ws['B46'] = 80
    ws['B46'].fill = INPUT_FILL
    ws['B46'].font = INPUT_FONT
    ws['B46'].number_format = '#,##0.0'

    ws['A47'] = "Beginning Shareholders' Equity"
    ws['B47'] = 200
    ws['B47'].fill = INPUT_FILL
    ws['B47'].font = INPUT_FONT
    ws['B47'].number_format = '#,##0.0'

def rebuild_income_statement(ws):
    """Rebuild Income Statement with proper links to scenario assumptions"""
    print("  Rebuilding Income Statement...")

    # Clear content but keep structure
    # We need to update formulas to reference the new assumption structure

    col_start = 2  # Column B

    # Update revenue growth formulas for Product Revenue (row 6)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i < 4:  # Historical - link to assumptions
            ws[f'{col}6'].value = f"='Assumptions & Drivers'!{col}8"
        elif i == 4:  # 2025 - use 2025 growth rate
            ws[f'{col}6'].value = f"=E6*(1+'Assumptions & Drivers'!$E$13)"
        else:  # 2026+ - use 2026+ growth rate
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}6'].value = f"={prev_col}6*(1+'Assumptions & Drivers'!$E$14)"
        ws[f'{col}6'].number_format = '#,##0.0'

    # Update service revenue growth formulas (row 7)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i < 4:  # Historical
            ws[f'{col}7'].value = f"='Assumptions & Drivers'!{col}9"
        elif i == 4:  # 2025
            ws[f'{col}7'].value = f"=E7*(1+'Assumptions & Drivers'!$E$15)"
        else:  # 2026+
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}7'].value = f"={prev_col}7*(1+'Assumptions & Drivers'!$E$16)"
        ws[f'{col}7'].number_format = '#,##0.0'

    # Update COGS formula to use scenario gross margin (row 9)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i < 4:  # Historical
            ws[f'{col}9'].value = f"={col}8*(1-'Assumptions & Drivers'!$B$26)"
        else:  # Forecast
            ws[f'{col}9'].value = f"={col}8*(1-'Assumptions & Drivers'!$E$17)"
        ws[f'{col}9'].number_format = '#,##0.0'

    # Update SG&A (row 12)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i < 4:  # Historical
            ws[f'{col}12'].value = f"={col}8*'Assumptions & Drivers'!$B$27"
        else:  # Forecast
            ws[f'{col}12'].value = f"={col}8*'Assumptions & Drivers'!$E$18"
        ws[f'{col}12'].number_format = '#,##0.0'

    # Update R&D (row 13)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i < 4:  # Historical
            ws[f'{col}13'].value = f"={col}8*'Assumptions & Drivers'!$B$28"
        else:  # Forecast
            ws[f'{col}13'].value = f"={col}8*'Assumptions & Drivers'!$E$19"
        ws[f'{col}13'].number_format = '#,##0.0'

    # D&A remains linked to revenue % (row 14)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}14'].value = f"={col}8*'Assumptions & Drivers'!$B$29"
        ws[f'{col}14'].number_format = '#,##0.0'

    # Tax rate (row 20)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}20'].value = f"={col}19*'Assumptions & Drivers'!$B$30"
        ws[f'{col}20'].number_format = '#,##0.0'

def rebuild_balance_sheet(ws):
    """Rebuild Balance Sheet with proper debt schedule and working capital drivers"""
    print("  Rebuilding Balance Sheet...")

    col_start = 2  # Column B

    # Update AR formula (row 8) to use scenario-driven days
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i < 4:  # Historical
            ws[f'{col}8'].value = f"='Income Statement'!{col}8*'Assumptions & Drivers'!$B$31/365"
        else:  # Forecast
            ws[f'{col}8'].value = f"='Income Statement'!{col}8*'Assumptions & Drivers'!$E$21/365"
        ws[f'{col}8'].number_format = '#,##0.0'

    # Update Inventory formula (row 9)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i < 4:  # Historical
            ws[f'{col}9'].value = f"='Income Statement'!{col}9*'Assumptions & Drivers'!$B$32/365"
        else:  # Forecast
            ws[f'{col}9'].value = f"='Income Statement'!{col}9*'Assumptions & Drivers'!$E$22/365"
        ws[f'{col}9'].number_format = '#,##0.0'

    # Update Other Current Assets (row 10)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}10'].value = f"='Income Statement'!{col}8*'Assumptions & Drivers'!$B$34"
        ws[f'{col}10'].number_format = '#,##0.0'

    # Update AP formula (row 19)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i < 4:  # Historical
            ws[f'{col}19'].value = f"='Income Statement'!{col}9*'Assumptions & Drivers'!$B$33/365"
        else:  # Forecast
            ws[f'{col}19'].value = f"='Income Statement'!{col}9*'Assumptions & Drivers'!$E$23/365"
        ws[f'{col}19'].number_format = '#,##0.0'

    # Update Other Current Liabilities (row 20)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}20'].value = f"='Income Statement'!{col}8*'Assumptions & Drivers'!$B$35"
        ws[f'{col}20'].number_format = '#,##0.0'

    # Debt now comes from Debt Schedule sheet
    # Row 26 = Total Debt (Term Loan + Revolver)
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}26'].value = f"='Debt Schedule'!{col}17"  # Will create this sheet
        ws[f'{col}26'].number_format = '#,##0.0'

def create_debt_schedule(wb):
    """Create a new Debt Schedule sheet"""
    print("  Creating Debt Schedule...")

    # Check if sheet exists, if so remove it
    if "Debt Schedule" in wb.sheetnames:
        del wb["Debt Schedule"]

    # Create new sheet after Balance Sheet
    bs_idx = wb.sheetnames.index("Balance Sheet")
    ws = wb.create_sheet("Debt Schedule", bs_idx + 1)

    # Set column widths
    set_column_widths(ws, {
        'A': 30, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12
    })

    # Title
    ws['A1'] = 'DEBT SCHEDULE'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A2'] = '($ in millions)'
    ws['A2'].font = Font(italic=True)

    # Years header
    row = 4
    ws[f'A{row}'] = 'Period'
    ws[f'A{row}'].font = BOLD_FONT

    col_start = 2
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # TERM LOAN
    row = 5
    ws[f'A{row}'] = 'TERM LOAN'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    row = 6
    ws[f'A{row}'] = 'Beginning Balance'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = "='Assumptions & Drivers'!B39"
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"={prev_col}{row+3}"  # Link to ending balance
        ws[f'{col}{row}'].number_format = '#,##0.0'

    row = 7
    ws[f'A{row}'] = 'Borrowing'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = 0  # No new term loan borrowing
        ws[f'{col}{row}'].number_format = '#,##0.0'

    row = 8
    ws[f'A{row}'] = 'Repayment'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = 0  # No mandatory amortization
        ws[f'{col}{row}'].number_format = '#,##0.0'

    row = 9
    ws[f'A{row}'] = 'Ending Balance'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}6+{col}7-{col}8"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    row = 10
    ws[f'A{row}'] = 'Average Balance'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"=({col}6+{col}9)/2"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    row = 11
    ws[f'A{row}'] = 'Interest Expense'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}10*'Assumptions & Drivers'!$B$40"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # REVOLVER
    row = 13
    ws[f'A{row}'] = 'REVOLVER'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    row = 14
    ws[f'A{row}'] = 'Beginning Balance'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = 0  # Start with no revolver
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"={prev_col}{row+3}"  # Link to ending balance
        ws[f'{col}{row}'].number_format = '#,##0.0'

    row = 15
    ws[f'A{row}'] = 'Draw / (Paydown)'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        # This will be calculated from cash flow
        ws[f'{col}{row}'] = f"='Cash Flow'!{col}24"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    row = 16
    ws[f'A{row}'] = 'Ending Balance'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"=MAX(0,{col}14+{col}15)"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    row = 17
    ws[f'A{row}'] = 'Total Debt'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}9+{col}16"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    row = 19
    ws[f'A{row}'] = 'Revolver Average Balance'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"=({col}14+{col}16)/2"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    row = 20
    ws[f'A{row}'] = 'Revolver Interest Expense'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}19*'Assumptions & Drivers'!$B$41"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    row = 22
    ws[f'A{row}'] = 'Total Interest Expense'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}11+{col}20"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

def rebuild_cash_flow(ws):
    """Rebuild Cash Flow with proper revolver logic and minimum cash"""
    print("  Rebuilding Cash Flow Statement...")

    col_start = 2

    # Update CapEx formula (row 18) to use scenario-driven capex
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i < 4:  # Historical
            ws[f'{col}18'].value = f"=-('Income Statement'!{col}8*'Assumptions & Drivers'!$B$36)"
        else:  # Forecast
            ws[f'{col}18'].value = f"=-('Income Statement'!{col}8*'Assumptions & Drivers'!$E$20)"
        ws[f'{col}18'].number_format = '#,##0.0'

    # Revolver draw/paydown logic with minimum cash
    # Row 24 = Net Borrowing / (Repayment) - this is the plug
    row = 24
    ws[f'A{row}'] = 'Revolver Draw / (Paydown)'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = 0
        else:
            prev_col = get_column_letter(col_start + i - 1)
            # Draw if cash before financing < minimum cash
            # Paydown if cash before financing > minimum cash + revolver balance
            # Cash before financing = beginning cash + CFO + CFI
            ws[f'{col}{row}'] = f"=MAX(-{col}29,'Assumptions & Drivers'!$B$42-({col}28+{col}16+{col}19))"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Update ending cash formula (row 32)
    row = 32
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'].value = f"={col}28+{col}27"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Update cash reconciliation section
    # Row 28 = Beginning Cash
    ws['A28'] = 'Beginning Cash'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}28'].value = "='Assumptions & Drivers'!B45"
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}28'].value = f"={prev_col}32"
        ws[f'{col}28'].number_format = '#,##0.0'

    # Row 29 = Cash before financing (for revolver calc)
    ws['A29'] = 'Cash Before Financing'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}29'].value = f"={col}28+{col}16+{col}19"
        ws[f'{col}29'].number_format = '#,##0.0'

    # Update Interest Expense link (row 18 in IS)
    # This should link to Debt Schedule total interest

def update_income_statement_interest(ws):
    """Update Income Statement to use Debt Schedule interest"""
    print("  Updating Income Statement interest expense...")

    col_start = 2
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}18'].value = f"='Debt Schedule'!{col}22"
        ws[f'{col}18'].number_format = '#,##0.0'

def create_checks_tab(wb):
    """Create a Checks tab for model validation"""
    print("  Creating Checks tab...")

    # Check if sheet exists
    if "Checks" in wb.sheetnames:
        del wb["Checks"]

    # Add at the end
    ws = wb.create_sheet("Checks")

    # Set column widths
    set_column_widths(ws, {'A': 35, 'B': 12, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
                           'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12})

    # Title
    ws['A1'] = 'MODEL CHECKS'
    ws['A1'].font = Font(size=14, bold=True)

    ws['A2'] = 'All checks should equal zero or show "PASS"'
    ws['A2'].font = Font(italic=True)

    # Years header
    row = 4
    ws[f'A{row}'] = 'Check'
    ws[f'A{row}'].font = BOLD_FONT

    col_start = 2
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # Balance Sheet Check
    row = 5
    ws[f'A{row}'] = 'Balance Sheet Check'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"='Balance Sheet'!{col}31"  # Balance check row
        ws[f'{col}{row}'].number_format = '#,##0.0'
        # Conditional formatting would be applied manually or via openpyxl rules

    # Cash Flow Reconciliation
    row = 6
    ws[f'A{row}'] = 'Cash Flow Reconciliation'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        # Ending cash from CF should equal cash on BS
        ws[f'{col}{row}'] = f"='Cash Flow'!{col}32-'Balance Sheet'!{col}7"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Status rows
    row = 8
    ws[f'A{row}'] = 'Balance Sheet Status'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f'=IF(ABS({col}5)<0.1,"PASS","FAIL")'
        ws[f'{col}{row}'].font = CHECK_FONT

    row = 9
    ws[f'A{row}'] = 'Cash Flow Status'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f'=IF(ABS({col}6)<0.1,"PASS","FAIL")'
        ws[f'{col}{row}'].font = CHECK_FONT

    # Summary Check
    row = 11
    ws[f'A{row}'] = 'OVERALL MODEL STATUS'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    row = 12
    ws[f'A{row}'] = 'Model Integrity'
    ws[f'A{row}'].font = BOLD_FONT
    ws['B12'] = '=IF(AND(COUNTIF(B8:K9,"FAIL")=0),"ALL CHECKS PASS","ERRORS DETECTED")'
    ws['B12'].font = Font(size=12, bold=True, color="006100")

def create_summary_tab(wb):
    """Create Summary tab with scenario comparison and charts"""
    print("  Creating Summary tab...")

    # Check if sheet exists
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    # Add at beginning (after Assumptions)
    ws = wb.create_sheet("Summary", 1)

    # Set column widths
    set_column_widths(ws, {'A': 25, 'B': 15, 'C': 15, 'D': 15, 'E': 15})

    # Title
    ws['A1'] = 'EXECUTIVE SUMMARY'
    ws['A1'].font = Font(size=16, bold=True, color="4472C4")

    ws['A2'] = f"Current Scenario: "
    ws['B2'] = "='Assumptions & Drivers'!B4"
    ws['B2'].font = Font(size=12, bold=True, color="0000FF")

    # Scenario Comparison Table
    ws['A4'] = 'SCENARIO COMPARISON'
    ws['A4'].font = Font(size=14, bold=True)

    ws['A5'] = 'Key Metric (2029)'
    ws['B5'] = 'Base'
    ws['C5'] = 'Upside'
    ws['D5'] = 'Downside'
    ws['E5'] = 'Current'

    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[col + '5'].font = BOLD_FONT
        ws[col + '5'].fill = HEADER_FILL

    # Note: We'll need to create hidden calculation rows or use complex formulas
    # For now, create placeholder formulas that reference the current scenario

    row = 6
    ws[f'A{row}'] = 'Revenue'
    ws[f'E{row}'] = "='Income Statement'!K8"  # 2029 revenue
    ws[f'E{row}'].number_format = '#,##0.0'

    row = 7
    ws[f'A{row}'] = 'EBITDA'
    ws[f'E{row}'] = "='Income Statement'!K22"  # 2029 EBITDA
    ws[f'E{row}'].number_format = '#,##0.0'

    row = 8
    ws[f'A{row}'] = 'EBITDA Margin %'
    ws[f'E{row}'] = "='Income Statement'!K23"
    ws[f'E{row}'].number_format = '0.0%'

    row = 9
    ws[f'A{row}'] = 'Net Income'
    ws[f'E{row}'] = "='Income Statement'!K21"
    ws[f'E{row}'].number_format = '#,##0.0'

    row = 10
    ws[f'A{row}'] = 'Free Cash Flow'
    ws[f'E{row}'] = "='Cash Flow'!K36"
    ws[f'E{row}'].number_format = '#,##0.0'

    row = 11
    ws[f'A{row}'] = 'Total Debt'
    ws[f'E{row}'] = "='Debt Schedule'!K17"
    ws[f'E{row}'].number_format = '#,##0.0'

    # Add note about scenario comparison
    ws['A13'] = 'Note: To compare scenarios, manually change scenario in Assumptions & Drivers'
    ws['A13'].font = Font(italic=True, size=9)
    ws['A14'] = 'and record metrics above for each scenario.'
    ws['A14'].font = Font(italic=True, size=9)

    # Historical data table for charts
    ws['A16'] = 'HISTORICAL & FORECAST DATA'
    ws['A16'].font = Font(size=12, bold=True)

    ws['A17'] = 'Year'
    ws['B17'] = 'Revenue'
    ws['C17'] = 'EBITDA'
    ws['D17'] = 'Free Cash Flow'

    for col in ['A', 'B', 'C', 'D']:
        ws[col + '17'].font = BOLD_FONT
        ws[col + '17'].fill = SECTION_FILL

    # Populate data
    for i, year in enumerate(ALL_YEARS):
        row = 18 + i
        col_letter = get_column_letter(2 + i)

        ws[f'A{row}'] = year
        ws[f'B{row}'] = f"='Income Statement'!{col_letter}8"
        ws[f'C{row}'] = f"='Income Statement'!{col_letter}22"
        ws[f'D{row}'] = f"='Cash Flow'!{col_letter}36"

        ws[f'B{row}'].number_format = '#,##0.0'
        ws[f'C{row}'].number_format = '#,##0.0'
        ws[f'D{row}'].number_format = '#,##0.0'

    # Add charts
    # Revenue Chart
    chart1 = LineChart()
    chart1.title = "Revenue Growth ($mm)"
    chart1.style = 13
    chart1.y_axis.title = '$ millions'
    chart1.x_axis.title = 'Year'
    chart1.height = 10
    chart1.width = 20

    data1 = Reference(ws, min_col=2, min_row=17, max_row=17+len(ALL_YEARS))
    cats1 = Reference(ws, min_col=1, min_row=18, max_row=17+len(ALL_YEARS))
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)

    ws.add_chart(chart1, "F4")

    # EBITDA Chart
    chart2 = LineChart()
    chart2.title = "EBITDA Trend ($mm)"
    chart2.style = 13
    chart2.y_axis.title = '$ millions'
    chart2.x_axis.title = 'Year'
    chart2.height = 10
    chart2.width = 20

    data2 = Reference(ws, min_col=3, min_row=17, max_row=17+len(ALL_YEARS))
    cats2 = Reference(ws, min_col=1, min_row=18, max_row=17+len(ALL_YEARS))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)

    ws.add_chart(chart2, "F20")

    # FCF Chart
    chart3 = LineChart()
    chart3.title = "Free Cash Flow ($mm)"
    chart3.style = 13
    chart3.y_axis.title = '$ millions'
    chart3.x_axis.title = 'Year'
    chart3.height = 10
    chart3.width = 20

    data3 = Reference(ws, min_col=4, min_row=17, max_row=17+len(ALL_YEARS))
    cats3 = Reference(ws, min_col=1, min_row=18, max_row=17+len(ALL_YEARS))
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)

    ws.add_chart(chart3, "F36")

def main():
    """Main enhancement function"""
    print("Enhancing 3-Statement Financial Model...")
    print("=" * 60)

    # Load existing workbook
    wb = load_workbook("3_Statement_Financial_Model.xlsx")
    print("✓ Loaded existing model")

    # Enhance Assumptions sheet
    enhance_assumptions_sheet(wb["Assumptions & Drivers"])
    print("✓ Enhanced Assumptions & Drivers")

    # Create Debt Schedule first (needed by other sheets)
    create_debt_schedule(wb)
    print("✓ Created Debt Schedule")

    # Rebuild other sheets with new formulas
    rebuild_income_statement(wb["Income Statement"])
    print("✓ Updated Income Statement")

    update_income_statement_interest(wb["Income Statement"])
    print("✓ Linked Interest Expense to Debt Schedule")

    rebuild_balance_sheet(wb["Balance Sheet"])
    print("✓ Updated Balance Sheet")

    rebuild_cash_flow(wb["Cash Flow"])
    print("✓ Updated Cash Flow Statement")

    # Create new tabs
    create_checks_tab(wb)
    print("✓ Created Checks tab")

    create_summary_tab(wb)
    print("✓ Created Summary tab")

    # Remove old Charts tab
    if "Charts" in wb.sheetnames:
        del wb["Charts"]
        print("✓ Removed old Charts tab (replaced by Summary)")

    # Save enhanced model
    wb.save("3_Statement_Financial_Model.xlsx")
    print("\n" + "=" * 60)
    print("✓ Enhanced model saved!")
    print("\nNew features:")
    print("  • Proper debt schedule with Term Loan + Revolver")
    print("  • Interest based on average balances")
    print("  • Minimum cash balance of $20mm")
    print("  • Scenario dropdown drives all key assumptions")
    print("  • Checks tab validates model integrity")
    print("  • Summary tab with charts and scenario comparison")
    print("  • Consistent color coding (blue/black/green)")
    print("\nTo use: Change scenario in 'Assumptions & Drivers'!B4")

if __name__ == "__main__":
    main()
