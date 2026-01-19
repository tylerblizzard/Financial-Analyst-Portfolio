#!/usr/bin/env python3
"""
Add DCF Valuation Tab to 3-Statement Model
Completes the IB-grade model with full valuation capabilities
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, BarChart, Reference

# Color definitions
INPUT_FILL = PatternFill(start_color="D6E4F5", end_color="D6E4F5", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
CHECK_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

INPUT_FONT = Font(color="0000FF", bold=True)
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)

# Years
HISTORICAL_YEARS = [2021, 2022, 2023, 2024]
FORECAST_YEARS = [2025, 2026, 2027, 2028, 2029]
ALL_YEARS = HISTORICAL_YEARS + FORECAST_YEARS

def set_column_widths(ws, widths):
    """Set column widths for a worksheet"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def create_dcf_tab(wb):
    """Create comprehensive DCF valuation tab"""
    print("  Creating DCF tab...")

    # Check if sheet exists
    if "DCF" in wb.sheetnames:
        del wb["DCF"]

    # Add after Cash Flow
    cf_idx = wb.sheetnames.index("Cash Flow")
    ws = wb.create_sheet("DCF", cf_idx + 1)

    # Set column widths
    set_column_widths(ws, {
        'A': 35, 'B': 4, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12
    })

    # Title
    ws['A1'] = 'DCF VALUATION ANALYSIS'
    ws['A1'].font = Font(size=16, bold=True, color="4472C4")
    ws['A2'] = 'Unlevered Free Cash Flow Method'
    ws['A2'].font = Font(size=12, italic=True)

    # Current scenario indicator
    ws['A3'] = 'Scenario:'
    ws['B3'] = "='Assumptions & Drivers'!B4"
    ws['B3'].font = Font(bold=True, color="0000FF")

    # SECTION 1: UNLEVERED FREE CASH FLOW
    row = 5
    ws[f'A{row}'] = 'UNLEVERED FREE CASH FLOW'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    # Year headers (forecast years only for DCF)
    row = 6
    ws[f'A{row}'] = 'Period'
    ws[f'A{row}'].font = BOLD_FONT

    col_start = 3  # Column C (skip B for spacing)
    for i, year in enumerate(FORECAST_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # EBIT
    row = 7
    ws[f'A{row}'] = 'EBIT'
    for i, year in enumerate(FORECAST_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(2 + 4 + i)  # F, G, H, I, J for forecast years
        ws[f'{col}{row}'] = f"='Income Statement'!{year_col}16"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Tax Rate
    row = 8
    ws[f'A{row}'] = 'Tax Rate'
    for i, year in enumerate(FORECAST_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(2 + 4 + i)
        ws[f'{col}{row}'] = f"='Assumptions & Drivers'!$B$30"
        ws[f'{col}{row}'].number_format = '0.0%'

    # NOPAT (EBIT * (1-Tax))
    row = 9
    ws[f'A{row}'] = 'NOPAT (EBIT × (1-Tax))'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(FORECAST_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}7*(1-{col}8)"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Plus: D&A
    row = 10
    ws[f'A{row}'] = 'Plus: D&A'
    for i, year in enumerate(FORECAST_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(2 + 4 + i)
        ws[f'{col}{row}'] = f"='Income Statement'!{year_col}14"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Less: CapEx
    row = 11
    ws[f'A{row}'] = 'Less: CapEx'
    for i, year in enumerate(FORECAST_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(2 + 4 + i)
        ws[f'{col}{row}'] = f"='Cash Flow'!{year_col}18"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Less: Change in NWC
    row = 12
    ws[f'A{row}'] = 'Less: Increase in NWC'
    for i, year in enumerate(FORECAST_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(2 + 4 + i)
        # Sum of working capital changes
        ws[f'{col}{row}'] = f"=-SUM('Cash Flow'!{year_col}9:{year_col}13)"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Unlevered FCF
    row = 13
    fcf_row = row
    ws[f'A{row}'] = 'Unlevered Free Cash Flow'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    for i, year in enumerate(FORECAST_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}9+{col}10+{col}11+{col}12"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # SECTION 2: WACC CALCULATION
    row = 16
    ws[f'A{row}'] = 'WACC CALCULATION'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    # Cost of Equity (CAPM)
    row = 17
    ws[f'A{row}'] = 'Cost of Equity (CAPM)'
    ws[f'A{row}'].font = BOLD_FONT

    row = 18
    ws[f'A{row}'] = 'Risk-Free Rate'
    ws['C18'] = 0.045
    ws['C18'].fill = INPUT_FILL
    ws['C18'].font = INPUT_FONT
    ws['C18'].number_format = '0.0%'

    row = 19
    ws[f'A{row}'] = 'Equity Risk Premium'
    ws['C19'] = 0.065
    ws['C19'].fill = INPUT_FILL
    ws['C19'].font = INPUT_FONT
    ws['C19'].number_format = '0.0%'

    row = 20
    ws[f'A{row}'] = 'Beta'
    ws['C20'] = 1.2
    ws['C20'].fill = INPUT_FILL
    ws['C20'].font = INPUT_FONT
    ws['C20'].number_format = '0.00'

    row = 21
    ws[f'A{row}'] = 'Cost of Equity'
    ws[f'A{row}'].font = BOLD_FONT
    ws['C21'] = '=C18+C20*C19'
    ws['C21'].number_format = '0.0%'
    ws['C21'].font = BOLD_FONT

    # Cost of Debt
    row = 23
    ws[f'A{row}'] = 'Cost of Debt'
    ws[f'A{row}'].font = BOLD_FONT

    row = 24
    ws[f'A{row}'] = 'Pre-Tax Cost of Debt'
    ws['C24'] = "=('Assumptions & Drivers'!B40+'Assumptions & Drivers'!B41)/2"
    ws['C24'].number_format = '0.0%'

    row = 25
    ws[f'A{row}'] = 'Tax Rate'
    ws['C25'] = "='Assumptions & Drivers'!B30"
    ws['C25'].number_format = '0.0%'

    row = 26
    ws[f'A{row}'] = 'After-Tax Cost of Debt'
    ws[f'A{row}'].font = BOLD_FONT
    ws['C26'] = '=C24*(1-C25)'
    ws['C26'].number_format = '0.0%'
    ws['C26'].font = BOLD_FONT

    # Capital Structure
    row = 28
    ws[f'A{row}'] = 'Target Capital Structure'
    ws[f'A{row}'].font = BOLD_FONT

    row = 29
    ws[f'A{row}'] = 'Target Debt %'
    ws['C29'] = 0.30
    ws['C29'].fill = INPUT_FILL
    ws['C29'].font = INPUT_FONT
    ws['C29'].number_format = '0.0%'

    row = 30
    ws[f'A{row}'] = 'Target Equity %'
    ws['C30'] = '=1-C29'
    ws['C30'].number_format = '0.0%'

    # WACC
    row = 32
    ws[f'A{row}'] = 'WACC'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    ws['C32'] = '=C21*C30+C26*C29'
    ws['C32'].number_format = '0.0%'
    ws['C32'].font = Font(bold=True, size=14, color="4472C4")

    # SECTION 3: TERMINAL VALUE
    row = 35
    ws[f'A{row}'] = 'TERMINAL VALUE CALCULATION'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    # Perpetuity Growth Method
    row = 36
    ws[f'A{row}'] = 'Method 1: Perpetuity Growth'
    ws[f'A{row}'].font = BOLD_FONT

    row = 37
    ws[f'A{row}'] = 'Terminal Year FCF'
    ws['C37'] = f"={get_column_letter(col_start + len(FORECAST_YEARS) - 1)}{fcf_row}"
    ws['C37'].number_format = '#,##0.0'

    row = 38
    ws[f'A{row}'] = 'Perpetual Growth Rate'
    ws['C38'] = 0.025
    ws['C38'].fill = INPUT_FILL
    ws['C38'].font = INPUT_FONT
    ws['C38'].number_format = '0.0%'

    row = 39
    ws[f'A{row}'] = 'Terminal Value (Perpetuity)'
    ws[f'A{row}'].font = BOLD_FONT
    ws['C39'] = '=C37*(1+C38)/(C32-C38)'
    ws['C39'].number_format = '#,##0.0'
    ws['C39'].font = BOLD_FONT

    # Exit Multiple Method
    row = 41
    ws[f'A{row}'] = 'Method 2: Exit Multiple'
    ws[f'A{row}'].font = BOLD_FONT

    row = 42
    ws[f'A{row}'] = 'Terminal Year EBITDA'
    last_year_col = get_column_letter(2 + len(ALL_YEARS) - 1)
    ws['C42'] = f"='Income Statement'!{last_year_col}22"
    ws['C42'].number_format = '#,##0.0'

    row = 43
    ws[f'A{row}'] = 'Exit EV/EBITDA Multiple'
    ws['C43'] = 8.5
    ws['C43'].fill = INPUT_FILL
    ws['C43'].font = INPUT_FONT
    ws['C43'].number_format = '0.0x'

    row = 44
    ws[f'A{row}'] = 'Terminal Value (Exit Multiple)'
    ws[f'A{row}'].font = BOLD_FONT
    ws['C44'] = '=C42*C43'
    ws['C44'].number_format = '#,##0.0'
    ws['C44'].font = BOLD_FONT

    # SECTION 4: DCF VALUATION
    row = 47
    ws[f'A{row}'] = 'DCF VALUATION - PERPETUITY GROWTH'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    # Discount periods
    row = 48
    ws[f'A{row}'] = 'Period'
    col_start_val = 3
    for i, year in enumerate(FORECAST_YEARS):
        col = get_column_letter(col_start_val + i)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = BOLD_FONT
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    row = 49
    ws[f'A{row}'] = 'Discount Period (years)'
    for i in range(len(FORECAST_YEARS)):
        col = get_column_letter(col_start_val + i)
        ws[f'{col}{row}'] = i + 1
        ws[f'{col}{row}'].number_format = '0.0'

    row = 50
    ws[f'A{row}'] = 'Unlevered FCF'
    for i in range(len(FORECAST_YEARS)):
        col = get_column_letter(col_start_val + i)
        ws[f'{col}{row}'] = f"={col}{fcf_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    row = 51
    ws[f'A{row}'] = 'Discount Factor'
    for i in range(len(FORECAST_YEARS)):
        col = get_column_letter(col_start_val + i)
        ws[f'{col}{row}'] = f"=1/(1+$C$32)^{col}49"
        ws[f'{col}{row}'].number_format = '0.000'

    row = 52
    ws[f'A{row}'] = 'PV of FCF'
    ws[f'A{row}'].font = BOLD_FONT
    for i in range(len(FORECAST_YEARS)):
        col = get_column_letter(col_start_val + i)
        ws[f'{col}{row}'] = f"={col}50*{col}51"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Terminal Value PV
    row = 54
    ws[f'A{row}'] = 'Terminal Value'
    last_forecast_col = get_column_letter(col_start_val + len(FORECAST_YEARS) - 1)
    ws[f'{last_forecast_col}{row}'] = '=C39'
    ws[f'{last_forecast_col}{row}'].number_format = '#,##0.0'

    row = 55
    ws[f'A{row}'] = 'PV of Terminal Value'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'{last_forecast_col}{row}'] = f"={last_forecast_col}54*{last_forecast_col}51"
    ws[f'{last_forecast_col}{row}'].number_format = '#,##0.0'
    ws[f'{last_forecast_col}{row}'].font = BOLD_FONT

    # Valuation Summary
    row = 57
    ws[f'A{row}'] = 'VALUATION SUMMARY (Perpetuity)'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    row = 58
    ws[f'A{row}'] = 'PV of Forecast FCFs'
    ws['C58'] = f"=SUM(C52:{last_forecast_col}52)"
    ws['C58'].number_format = '#,##0.0'

    row = 59
    ws[f'A{row}'] = 'PV of Terminal Value'
    ws['C59'] = f"={last_forecast_col}55"
    ws['C59'].number_format = '#,##0.0'

    row = 60
    ws[f'A{row}'] = 'Enterprise Value'
    ws[f'A{row}'].font = BOLD_FONT
    ws['C60'] = '=C58+C59'
    ws['C60'].number_format = '#,##0.0'
    ws['C60'].font = BOLD_FONT

    # Net Debt Bridge
    row = 62
    ws[f'A{row}'] = 'Less: Net Debt'
    last_year_col = get_column_letter(2 + len(ALL_YEARS) - 1)
    ws['C62'] = f"='Debt Schedule'!{last_year_col}17-'Balance Sheet'!{last_year_col}7"
    ws['C62'].number_format = '#,##0.0'

    row = 63
    ws[f'A{row}'] = 'Equity Value'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    ws['C63'] = '=C60-C62'
    ws['C63'].number_format = '#,##0.0'
    ws['C63'].font = Font(bold=True, size=12, color="4472C4")

    # Per Share
    row = 65
    ws[f'A{row}'] = 'Fully Diluted Shares (mm)'
    ws['C65'] = 100
    ws['C65'].fill = INPUT_FILL
    ws['C65'].font = INPUT_FONT
    ws['C65'].number_format = '#,##0.0'

    row = 66
    ws[f'A{row}'] = 'Equity Value Per Share'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    ws['C66'] = '=C63/C65'
    ws['C66'].number_format = '$#,##0.00'
    ws['C66'].font = Font(bold=True, size=14, color="4472C4")

    # EXIT MULTIPLE VALUATION
    row = 69
    ws[f'A{row}'] = 'DCF VALUATION - EXIT MULTIPLE'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    row = 70
    ws[f'A{row}'] = 'PV of Forecast FCFs'
    ws['C70'] = '=C58'
    ws['C70'].number_format = '#,##0.0'

    row = 71
    ws[f'A{row}'] = 'Terminal Value (Exit Multiple)'
    ws['C71'] = '=C44'
    ws['C71'].number_format = '#,##0.0'

    row = 72
    ws[f'A{row}'] = 'PV of Terminal Value'
    ws['C72'] = f"=C71*{last_forecast_col}51"
    ws['C72'].number_format = '#,##0.0'

    row = 73
    ws[f'A{row}'] = 'Enterprise Value'
    ws[f'A{row}'].font = BOLD_FONT
    ws['C73'] = '=C70+C72'
    ws['C73'].number_format = '#,##0.0'
    ws['C73'].font = BOLD_FONT

    row = 74
    ws[f'A{row}'] = 'Less: Net Debt'
    ws['C74'] = '=C62'
    ws['C74'].number_format = '#,##0.0'

    row = 75
    ws[f'A{row}'] = 'Equity Value'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    ws['C75'] = '=C73-C74'
    ws['C75'].number_format = '#,##0.0'
    ws['C75'].font = Font(bold=True, size=12, color="4472C4")

    row = 76
    ws[f'A{row}'] = 'Equity Value Per Share'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    ws['C76'] = '=C75/C65'
    ws['C76'].number_format = '$#,##0.00'
    ws['C76'].font = Font(bold=True, size=14, color="4472C4")

    # SENSITIVITY TABLES
    row = 79
    ws[f'A{row}'] = 'SENSITIVITY ANALYSIS'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    row = 80
    ws[f'A{row}'] = 'Equity Value per Share Sensitivity (Perpetuity Method)'
    ws[f'A{row}'].font = BOLD_FONT

    # Create 2-way table: WACC vs Terminal Growth
    # Headers
    row = 81
    ws[f'A{row}'] = 'WACC / Growth Rate'
    ws[f'A{row}'].font = BOLD_FONT

    # Growth rates across top
    growth_rates = [0.015, 0.020, 0.025, 0.030, 0.035]
    for i, rate in enumerate(growth_rates):
        col = get_column_letter(3 + i)  # C, D, E, F, G
        ws[f'{col}{row}'] = rate
        ws[f'{col}{row}'].number_format = '0.0%'
        ws[f'{col}{row}'].font = BOLD_FONT
        ws[f'{col}{row}'].fill = SECTION_FILL

    # WACC down the side
    wacc_rates = [0.08, 0.09, 0.10, 0.11, 0.12]
    for i, rate in enumerate(wacc_rates):
        row = 82 + i
        ws[f'A{row}'] = rate
        ws[f'A{row}'].number_format = '0.0%'
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'A{row}'].fill = SECTION_FILL

    # Add note about data table
    ws['A88'] = 'Note: Use Excel Data Table feature (Data > What-If Analysis > Data Table)'
    ws['A88'].font = Font(italic=True, size=9)
    ws['A89'] = 'Row input: C38 (Growth Rate), Column input: C32 (WACC)'
    ws['A89'].font = Font(italic=True, size=9)

    # Add second sensitivity table for Exit Multiple method
    row = 92
    ws[f'A{row}'] = 'Equity Value per Share Sensitivity (Exit Multiple Method)'
    ws[f'A{row}'].font = BOLD_FONT

    row = 93
    ws[f'A{row}'] = 'WACC / Exit Multiple'
    ws[f'A{row}'].font = BOLD_FONT

    # Exit multiples across top
    exit_multiples = [7.0, 7.5, 8.0, 8.5, 9.0, 9.5, 10.0]
    for i, mult in enumerate(exit_multiples):
        col = get_column_letter(3 + i)
        ws[f'{col}{row}'] = mult
        ws[f'{col}{row}'].number_format = '0.0x'
        ws[f'{col}{row}'].font = BOLD_FONT
        ws[f'{col}{row}'].fill = SECTION_FILL

    # WACC down the side
    for i, rate in enumerate(wacc_rates):
        row = 94 + i
        ws[f'A{row}'] = rate
        ws[f'A{row}'].number_format = '0.0%'
        ws[f'A{row}'].font = BOLD_FONT
        ws[f'A{row}'].fill = SECTION_FILL

    ws['A100'] = 'Note: Use Excel Data Table feature (Data > What-If Analysis > Data Table)'
    ws['A100'].font = Font(italic=True, size=9)
    ws['A101'] = 'Row input: C43 (Exit Multiple), Column input: C32 (WACC)'
    ws['A101'].font = Font(italic=True, size=9)

def enhance_summary_with_valuation(ws):
    """Add DCF valuation outputs to Summary tab"""
    print("  Enhancing Summary tab with valuation...")

    # Add valuation section after existing content
    row = 30
    ws[f'A{row}'] = 'DCF VALUATION SUMMARY'
    ws[f'A{row}'].font = Font(size=14, bold=True)

    row = 32
    ws[f'A{row}'] = 'Method'
    ws[f'B{row}'] = 'Enterprise Value'
    ws[f'C{row}'] = 'Equity Value'
    ws[f'D{row}'] = 'Value per Share'

    for col in ['A', 'B', 'C', 'D']:
        ws[col + str(row)].font = BOLD_FONT
        ws[col + str(row)].fill = HEADER_FILL

    # Perpetuity Growth Method
    row = 33
    ws[f'A{row}'] = 'DCF - Perpetuity Growth'
    ws[f'B{row}'] = '=DCF!C60'
    ws[f'C{row}'] = '=DCF!C63'
    ws[f'D{row}'] = '=DCF!C66'

    ws[f'B{row}'].number_format = '#,##0.0'
    ws[f'C{row}'].number_format = '#,##0.0'
    ws[f'D{row}'].number_format = '$#,##0.00'

    # Exit Multiple Method
    row = 34
    ws[f'A{row}'] = 'DCF - Exit Multiple'
    ws[f'B{row}'] = '=DCF!C73'
    ws[f'C{row}'] = '=DCF!C75'
    ws[f'D{row}'] = '=DCF!C76'

    ws[f'B{row}'].number_format = '#,##0.0'
    ws[f'C{row}'].number_format = '#,##0.0'
    ws[f'D{row}'].number_format = '$#,##0.00'

    # Midpoint
    row = 35
    ws[f'A{row}'] = 'Midpoint'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'B{row}'] = '=(B33+B34)/2'
    ws[f'C{row}'] = '=(C33+C34)/2'
    ws[f'D{row}'] = '=(D33+D34)/2'

    ws[f'B{row}'].number_format = '#,##0.0'
    ws[f'C{row}'].number_format = '#,##0.0'
    ws[f'D{row}'].number_format = '$#,##0.00'
    ws[f'B{row}'].font = BOLD_FONT
    ws[f'C{row}'].font = BOLD_FONT
    ws[f'D{row}'].font = BOLD_FONT

    # Key assumptions
    row = 37
    ws[f'A{row}'] = 'KEY ASSUMPTIONS'
    ws[f'A{row}'].font = Font(size=12, bold=True)

    row = 38
    ws[f'A{row}'] = 'WACC'
    ws[f'B{row}'] = '=DCF!C32'
    ws[f'B{row}'].number_format = '0.0%'

    row = 39
    ws[f'A{row}'] = 'Terminal Growth Rate'
    ws[f'B{row}'] = '=DCF!C38'
    ws[f'B{row}'].number_format = '0.0%'

    row = 40
    ws[f'A{row}'] = 'Exit EBITDA Multiple'
    ws[f'B{row}'] = '=DCF!C43'
    ws[f'B{row}'].number_format = '0.0x'

    row = 41
    ws[f'A{row}'] = 'Terminal Year EBITDA'
    ws[f'B{row}'] = '=DCF!C42'
    ws[f'B{row}'].number_format = '#,##0.0'

def enhance_checks_with_dcf(ws):
    """Add DCF validation checks"""
    print("  Enhancing Checks tab with DCF validation...")

    # Add DCF checks below existing checks
    row = 13
    ws[f'A{row}'] = 'DCF CHECKS'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    row = 14
    ws[f'A{row}'] = 'WACC Check'
    ws[f'A{row}'].font = BOLD_FONT
    ws['B14'] = '=DCF!C32'
    ws['B14'].number_format = '0.0%'
    ws['C14'] = '=IF(DCF!C32>0.05,IF(DCF!C32<0.20,"PASS","FAIL"),"FAIL")'
    ws['C14'].font = Font(bold=True, color="006100")

    row = 15
    ws[f'A{row}'] = 'Terminal Growth < WACC'
    ws[f'A{row}'].font = BOLD_FONT
    ws['B15'] = '=DCF!C38'
    ws['B15'].number_format = '0.0%'
    ws['C15'] = '=IF(DCF!C38<DCF!C32,"PASS","FAIL")'
    ws['C15'].font = Font(bold=True, color="006100")

    row = 16
    ws[f'A{row}'] = 'Enterprise Value > 0'
    ws[f'A{row}'].font = BOLD_FONT
    ws['B16'] = '=DCF!C60'
    ws['B16'].number_format = '#,##0.0'
    ws['C16'] = '=IF(DCF!C60>0,"PASS","FAIL")'
    ws['C16'].font = Font(bold=True, color="006100")

    row = 17
    ws[f'A{row}'] = 'Equity Value > 0'
    ws[f'A{row}'].font = BOLD_FONT
    ws['B17'] = '=DCF!C63'
    ws['B17'].number_format = '#,##0.0'
    ws['C17'] = '=IF(DCF!C63>0,"PASS","FAIL")'
    ws['C17'].font = Font(bold=True, color="006100")

    # Update overall check
    ws['B12'] = '=IF(AND(COUNTIF(B8:K9,"FAIL")=0,COUNTIF(C14:C17,"FAIL")=0),"ALL CHECKS PASS","ERRORS DETECTED")'

def add_valuation_assumptions_section(ws):
    """Add DCF/Valuation assumptions to Assumptions & Drivers"""
    print("  Adding valuation assumptions to Assumptions & Drivers...")

    # Add at the end of the sheet
    row = 49
    ws[f'A{row}'] = 'DCF / VALUATION ASSUMPTIONS'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    row = 50
    ws[f'A{row}'] = 'Fully Diluted Shares (mm)'
    ws['B50'] = 100
    ws['B50'].fill = INPUT_FILL
    ws['B50'].font = INPUT_FONT
    ws['B50'].number_format = '#,##0.0'

    row = 51
    ws[f'A{row}'] = 'Risk-Free Rate'
    ws['B51'] = 0.045
    ws['B51'].fill = INPUT_FILL
    ws['B51'].font = INPUT_FONT
    ws['B51'].number_format = '0.0%'

    row = 52
    ws[f'A{row}'] = 'Equity Risk Premium'
    ws['B52'] = 0.065
    ws['B52'].fill = INPUT_FILL
    ws['B52'].font = INPUT_FONT
    ws['B52'].number_format = '0.0%'

    row = 53
    ws[f'A{row}'] = 'Beta'
    ws['B53'] = 1.2
    ws['B53'].fill = INPUT_FILL
    ws['B53'].font = INPUT_FONT
    ws['B53'].number_format = '0.00'

    row = 54
    ws[f'A{row}'] = 'Target Debt %'
    ws['B54'] = 0.30
    ws['B54'].fill = INPUT_FILL
    ws['B54'].font = INPUT_FONT
    ws['B54'].number_format = '0.0%'

    row = 55
    ws[f'A{row}'] = 'Terminal Growth Rate'
    ws['B55'] = 0.025
    ws['B55'].fill = INPUT_FILL
    ws['B55'].font = INPUT_FONT
    ws['B55'].number_format = '0.0%'

    row = 56
    ws[f'A{row}'] = 'Exit EV/EBITDA Multiple'
    ws['B56'] = 8.5
    ws['B56'].fill = INPUT_FILL
    ws['B56'].font = INPUT_FONT
    ws['B56'].number_format = '0.0x'

def main():
    """Main enhancement function"""
    print("Adding DCF Valuation to Financial Model...")
    print("=" * 60)

    # Load existing workbook
    wb = load_workbook("3_Statement_Financial_Model.xlsx")
    print("✓ Loaded existing model")

    # Create DCF tab
    create_dcf_tab(wb)
    print("✓ Created DCF tab")

    # Enhance Summary
    enhance_summary_with_valuation(wb["Summary"])
    print("✓ Enhanced Summary with valuation outputs")

    # Enhance Checks
    enhance_checks_with_dcf(wb["Checks"])
    print("✓ Enhanced Checks with DCF validation")

    # Add valuation assumptions
    add_valuation_assumptions_section(wb["Assumptions & Drivers"])
    print("✓ Added valuation assumptions")

    # Save enhanced model
    wb.save("3_Statement_Financial_Model.xlsx")
    print("\n" + "=" * 60)
    print("✓ Complete IB-grade model with DCF saved!")
    print("\nDCF Features Added:")
    print("  • Unlevered Free Cash Flow calculation")
    print("  • WACC calculation (CAPM + after-tax cost of debt)")
    print("  • Terminal value - Perpetuity Growth method")
    print("  • Terminal value - Exit Multiple method")
    print("  • NPV of all cash flows")
    print("  • Enterprise Value → Equity Value bridge")
    print("  • Equity Value per Share")
    print("  • Sensitivity table frameworks (WACC vs Growth, WACC vs Multiple)")
    print("  • Valuation summary on Summary tab")
    print("  • DCF integrity checks")
    print("\nModel is now complete and ready for IB use!")

if __name__ == "__main__":
    main()
