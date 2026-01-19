#!/usr/bin/env python3
"""
3-Statement Financial Model Builder
Creates an integrated Income Statement, Balance Sheet, and Cash Flow Statement
with IB-standard drivers and formatting
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation

# Color definitions
INPUT_FILL = PatternFill(start_color="D6E4F5", end_color="D6E4F5", fill_type="solid")  # Light blue
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Dark blue
SECTION_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Light gray
FORMULA_FONT = Font(color="000000")  # Black
INPUT_FONT = Font(color="0000FF", bold=True)  # Blue
HEADER_FONT = Font(color="FFFFFF", bold=True)  # White
BOLD_FONT = Font(bold=True)

# Border styles
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Years
HISTORICAL_YEARS = [2021, 2022, 2023, 2024]
FORECAST_YEARS = [2025, 2026, 2027, 2028, 2029]
ALL_YEARS = HISTORICAL_YEARS + FORECAST_YEARS

# Scenarios
SCENARIOS = ['Base', 'Upside', 'Downside']

def create_workbook():
    """Create and return a new workbook with all sheets"""
    wb = Workbook()

    # Remove default sheet and create our sheets in order
    wb.remove(wb.active)
    wb.create_sheet("Assumptions & Drivers", 0)
    wb.create_sheet("Income Statement", 1)
    wb.create_sheet("Balance Sheet", 2)
    wb.create_sheet("Cash Flow", 3)
    wb.create_sheet("Charts", 4)

    return wb

def set_column_widths(ws, widths):
    """Set column widths for a worksheet"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def apply_cell_style(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    """Apply styling to a cell"""
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format

def build_assumptions_sheet(wb):
    """Build the Assumptions & Drivers sheet"""
    ws = wb["Assumptions & Drivers"]

    # Set column widths
    set_column_widths(ws, {
        'A': 30, 'B': 15, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12
    })

    # Title
    ws['A1'] = '3-STATEMENT FINANCIAL MODEL'
    ws['A1'].font = Font(size=16, bold=True, color="4472C4")

    ws['A2'] = 'ASSUMPTIONS & DRIVERS'
    ws['A2'].font = Font(size=14, bold=True)

    # Scenario selector
    ws['A4'] = 'Scenario Selection:'
    ws['A4'].font = BOLD_FONT
    ws['B4'] = 'Base'
    ws['B4'].fill = INPUT_FILL
    ws['B4'].font = INPUT_FONT

    # Add data validation for scenario dropdown
    dv = DataValidation(type="list", formula1='"Base,Upside,Downside"', allow_blank=False)
    dv.add(ws['B4'])
    ws.add_data_validation(dv)

    # Years header
    row = 6
    ws[f'A{row}'] = 'INPUTS (in $mm)'
    ws[f'A{row}'].font = HEADER_FONT
    ws[f'A{row}'].fill = HEADER_FILL

    # Year headers
    col_start = 2  # Column B
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # Revenue Assumptions
    row += 1
    ws[f'A{row}'] = 'REVENUE ASSUMPTIONS'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # Product Revenue
    row += 1
    ws[f'A{row}'] = 'Product Revenue'
    ws['B' + str(row)] = 100
    ws['C' + str(row)] = 110
    ws['D' + str(row)] = 125
    ws['E' + str(row)] = 140
    for col in ['B', 'C', 'D', 'E']:
        ws[col + str(row)].fill = INPUT_FILL
        ws[col + str(row)].font = INPUT_FONT
        ws[col + str(row)].number_format = '#,##0.0'

    # Service Revenue
    row += 1
    ws[f'A{row}'] = 'Service Revenue'
    ws['B' + str(row)] = 50
    ws['C' + str(row)] = 55
    ws['D' + str(row)] = 62
    ws['E' + str(row)] = 70
    for col in ['B', 'C', 'D', 'E']:
        ws[col + str(row)].fill = INPUT_FILL
        ws[col + str(row)].font = INPUT_FONT
        ws[col + str(row)].number_format = '#,##0.0'

    # Product Revenue Growth %
    row += 1
    ws[f'A{row}'] = 'Product Revenue Growth %'
    for i, col_letter in enumerate([get_column_letter(col_start + 4 + i) for i in range(5)]):
        if col_letter == 'F':  # 2025
            ws[f'{col_letter}{row}'] = '=IF($B$4="Base",0.08,IF($B$4="Upside",0.12,0.05))'
        else:
            ws[f'{col_letter}{row}'] = '=IF($B$4="Base",0.06,IF($B$4="Upside",0.10,0.03))'
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # Service Revenue Growth %
    row += 1
    ws[f'A{row}'] = 'Service Revenue Growth %'
    for i, col_letter in enumerate([get_column_letter(col_start + 4 + i) for i in range(5)]):
        if col_letter == 'F':  # 2025
            ws[f'{col_letter}{row}'] = '=IF($B$4="Base",0.10,IF($B$4="Upside",0.15,0.07))'
        else:
            ws[f'{col_letter}{row}'] = '=IF($B$4="Base",0.08,IF($B$4="Upside",0.12,0.05))'
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # Operating Assumptions
    row += 2
    ws[f'A{row}'] = 'OPERATING ASSUMPTIONS'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # Gross Margin %
    row += 1
    ws[f'A{row}'] = 'Gross Margin %'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        if i < 4:  # Historical
            ws[f'{col_letter}{row}'] = 0.65
        else:  # Forecast
            ws[f'{col_letter}{row}'] = '=IF($B$4="Base",0.67,IF($B$4="Upside",0.70,0.64))'
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # SG&A % of Revenue
    row += 1
    ws[f'A{row}'] = 'SG&A % of Revenue'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        if i < 4:  # Historical
            ws[f'{col_letter}{row}'] = 0.25
        else:  # Forecast
            ws[f'{col_letter}{row}'] = '=IF($B$4="Base",0.24,IF($B$4="Upside",0.23,0.26))'
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # R&D % of Revenue
    row += 1
    ws[f'A{row}'] = 'R&D % of Revenue'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        if i < 4:  # Historical
            ws[f'{col_letter}{row}'] = 0.12
        else:  # Forecast
            ws[f'{col_letter}{row}'] = '=IF($B$4="Base",0.12,IF($B$4="Upside",0.13,0.10))'
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # D&A % of Revenue
    row += 1
    ws[f'A{row}'] = 'D&A % of Revenue'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        ws[f'{col_letter}{row}'] = 0.05
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # Tax Rate
    row += 1
    ws[f'A{row}'] = 'Tax Rate'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        ws[f'{col_letter}{row}'] = 0.25
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # Balance Sheet Assumptions
    row += 2
    ws[f'A{row}'] = 'BALANCE SHEET ASSUMPTIONS'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # AR Days
    row += 1
    ws[f'A{row}'] = 'AR Days'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        ws[f'{col_letter}{row}'] = 45
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0'

    # Inventory Days
    row += 1
    ws[f'A{row}'] = 'Inventory Days'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        ws[f'{col_letter}{row}'] = 60
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0'

    # AP Days
    row += 1
    ws[f'A{row}'] = 'AP Days'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        ws[f'{col_letter}{row}'] = 30
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0'

    # Other Current Assets % of Revenue
    row += 1
    ws[f'A{row}'] = 'Other Current Assets % of Rev'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        ws[f'{col_letter}{row}'] = 0.03
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # Other Current Liabilities % of Revenue
    row += 1
    ws[f'A{row}'] = 'Other Current Liab % of Rev'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        ws[f'{col_letter}{row}'] = 0.02
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # CapEx % of Revenue
    row += 1
    ws[f'A{row}'] = 'CapEx % of Revenue'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        if i < 4:  # Historical
            ws[f'{col_letter}{row}'] = 0.08
        else:  # Forecast
            ws[f'{col_letter}{row}'] = '=IF($B$4="Base",0.07,IF($B$4="Upside",0.09,0.06))'
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # Debt Assumptions
    row += 2
    ws[f'A{row}'] = 'DEBT ASSUMPTIONS'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # Beginning Debt
    row += 1
    ws[f'A{row}'] = 'Beginning Debt (2021)'
    ws['B' + str(row)] = 50
    ws['B' + str(row)].fill = INPUT_FILL
    ws['B' + str(row)].font = INPUT_FONT
    ws['B' + str(row)].number_format = '#,##0.0'

    # Revolver Interest Rate
    row += 1
    ws[f'A{row}'] = 'Revolver Interest Rate'
    for i, col_letter in enumerate([get_column_letter(col_start + i) for i in range(len(ALL_YEARS))]):
        ws[f'{col_letter}{row}'] = 0.05
        ws[f'{col_letter}{row}'].fill = INPUT_FILL
        ws[f'{col_letter}{row}'].font = INPUT_FONT
        ws[f'{col_letter}{row}'].number_format = '0.0%'

    # Beginning Equity
    row += 1
    ws[f'A{row}'] = 'Beginning Equity (2021)'
    ws['B' + str(row)] = 200
    ws['B' + str(row)].fill = INPUT_FILL
    ws['B' + str(row)].font = INPUT_FONT
    ws['B' + str(row)].number_format = '#,##0.0'

    # Beginning Cash
    row += 1
    ws[f'A{row}'] = 'Beginning Cash (2021)'
    ws['B' + str(row)] = 30
    ws['B' + str(row)].fill = INPUT_FILL
    ws['B' + str(row)].font = INPUT_FONT
    ws['B' + str(row)].number_format = '#,##0.0'

    # Beginning PP&E
    row += 1
    ws[f'A{row}'] = 'Beginning PP&E (2021)'
    ws['B' + str(row)] = 80
    ws['B' + str(row)].fill = INPUT_FILL
    ws['B' + str(row)].font = INPUT_FONT
    ws['B' + str(row)].number_format = '#,##0.0'

def build_income_statement(wb):
    """Build the Income Statement"""
    ws = wb["Income Statement"]

    # Set column widths
    set_column_widths(ws, {
        'A': 30, 'B': 15, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12
    })

    # Title
    ws['A1'] = 'INCOME STATEMENT'
    ws['A1'].font = Font(size=14, bold=True)

    ws['A2'] = '($ in millions)'
    ws['A2'].font = Font(italic=True)

    # Years header
    row = 4
    ws[f'A{row}'] = 'Period'
    ws[f'A{row}'].font = BOLD_FONT

    col_start = 2  # Column B
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # Revenue section
    row += 1
    ws[f'A{row}'] = 'Revenue'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # Product Revenue
    row += 1
    prod_rev_row = row
    ws[f'A{row}'] = 'Product Revenue'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)

        if i < 4:  # Historical - link to assumptions
            ws[f'{col}{row}'] = f"='Assumptions & Drivers'!{year_col}8"
        else:  # Forecast - calculate from growth rate
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"={prev_col}{row}*(1+'Assumptions & Drivers'!{year_col}10)"

        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Service Revenue
    row += 1
    svc_rev_row = row
    ws[f'A{row}'] = 'Service Revenue'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)

        if i < 4:  # Historical - link to assumptions
            ws[f'{col}{row}'] = f"='Assumptions & Drivers'!{year_col}9"
        else:  # Forecast - calculate from growth rate
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"={prev_col}{row}*(1+'Assumptions & Drivers'!{year_col}11)"

        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Total Revenue
    row += 1
    total_rev_row = row
    ws[f'A{row}'] = 'Total Revenue'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"=SUM({col}{prod_rev_row}:{col}{svc_rev_row})"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # COGS
    row += 1
    ws[f'A{row}'] = 'Cost of Goods Sold'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{total_rev_row}*(1-'Assumptions & Drivers'!{year_col}14)"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Gross Profit
    row += 1
    gross_profit_row = row
    ws[f'A{row}'] = 'Gross Profit'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{total_rev_row}-{col}{row-1}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Operating Expenses
    row += 1
    ws[f'A{row}'] = 'Operating Expenses'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # SG&A
    row += 1
    sga_row = row
    ws[f'A{row}'] = 'SG&A'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{total_rev_row}*'Assumptions & Drivers'!{year_col}15"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # R&D
    row += 1
    rnd_row = row
    ws[f'A{row}'] = 'R&D'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{total_rev_row}*'Assumptions & Drivers'!{year_col}16"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # D&A
    row += 1
    da_row = row
    ws[f'A{row}'] = 'D&A'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{total_rev_row}*'Assumptions & Drivers'!{year_col}17"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Total Operating Expenses
    row += 1
    total_opex_row = row
    ws[f'A{row}'] = 'Total Operating Expenses'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"=SUM({col}{sga_row}:{col}{da_row})"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # EBIT
    row += 1
    ebit_row = row
    ws[f'A{row}'] = 'EBIT'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{gross_profit_row}-{col}{total_opex_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Interest Expense
    row += 1
    interest_row = row
    ws[f'A{row}'] = 'Interest Expense'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        # Will link to BS debt * interest rate
        ws[f'{col}{row}'] = f"='Balance Sheet'!{col}26*'Assumptions & Drivers'!{year_col}31"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # EBT
    row += 1
    ebt_row = row
    ws[f'A{row}'] = 'EBT'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{ebit_row}-{col}{interest_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Taxes
    row += 1
    tax_row = row
    ws[f'A{row}'] = 'Taxes'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{ebt_row}*'Assumptions & Drivers'!{year_col}18"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Net Income
    row += 1
    ni_row = row
    ws[f'A{row}'] = 'Net Income'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{ebt_row}-{col}{tax_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Key metrics
    row += 2
    ws[f'A{row}'] = 'KEY METRICS'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # EBITDA
    row += 1
    ebitda_row = row
    ws[f'A{row}'] = 'EBITDA'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{ebit_row}+{col}{da_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # EBITDA Margin
    row += 1
    ws[f'A{row}'] = 'EBITDA Margin'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{ebitda_row}/{col}{total_rev_row}"
        ws[f'{col}{row}'].number_format = '0.0%'

    # Net Income Margin
    row += 1
    ws[f'A{row}'] = 'Net Income Margin'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{ni_row}/{col}{total_rev_row}"
        ws[f'{col}{row}'].number_format = '0.0%'

def build_balance_sheet(wb):
    """Build the Balance Sheet"""
    ws = wb["Balance Sheet"]

    # Set column widths
    set_column_widths(ws, {
        'A': 30, 'B': 15, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12
    })

    # Title
    ws['A1'] = 'BALANCE SHEET'
    ws['A1'].font = Font(size=14, bold=True)

    ws['A2'] = '($ in millions)'
    ws['A2'].font = Font(italic=True)

    # Years header
    row = 4
    ws[f'A{row}'] = 'Period Ending'
    ws[f'A{row}'].font = BOLD_FONT

    col_start = 2  # Column B
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # ASSETS
    row += 1
    ws[f'A{row}'] = 'ASSETS'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # Current Assets
    row += 1
    ws[f'A{row}'] = 'Current Assets'
    ws[f'A{row}'].font = BOLD_FONT

    # Cash
    row += 1
    cash_row = row
    ws[f'A{row}'] = 'Cash'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:  # 2021
            ws[f'{col}{row}'] = "='Assumptions & Drivers'!B33"
        else:
            # Link to CF ending cash
            ws[f'{col}{row}'] = f"='Cash Flow'!{col}32"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Accounts Receivable
    row += 1
    ar_row = row
    ws[f'A{row}'] = 'Accounts Receivable'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        # AR = Revenue * AR Days / 365
        ws[f'{col}{row}'] = f"='Income Statement'!{col}7*'Assumptions & Drivers'!{year_col}21/365"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Inventory
    row += 1
    inv_row = row
    ws[f'A{row}'] = 'Inventory'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        # Inventory = COGS * Inventory Days / 365
        ws[f'{col}{row}'] = f"='Income Statement'!{col}8*'Assumptions & Drivers'!{year_col}22/365"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Other Current Assets
    row += 1
    other_ca_row = row
    ws[f'A{row}'] = 'Other Current Assets'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"='Income Statement'!{col}7*'Assumptions & Drivers'!{year_col}24"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Total Current Assets
    row += 1
    total_ca_row = row
    ws[f'A{row}'] = 'Total Current Assets'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"=SUM({col}{cash_row}:{col}{other_ca_row})"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # PP&E
    row += 1
    ws[f'A{row}'] = 'PP&E, Net'
    ppe_row = row
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        if i == 0:  # 2021
            ws[f'{col}{row}'] = "='Assumptions & Drivers'!B34"
        else:
            # PP&E = Prior PP&E + CapEx - D&A
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"={prev_col}{row}+'Cash Flow'!{col}18-'Income Statement'!{col}13"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Total Assets
    row += 1
    total_assets_row = row
    ws[f'A{row}'] = 'Total Assets'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{total_ca_row}+{col}{ppe_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # LIABILITIES & EQUITY
    row += 2
    ws[f'A{row}'] = 'LIABILITIES & EQUITY'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # Current Liabilities
    row += 1
    ws[f'A{row}'] = 'Current Liabilities'
    ws[f'A{row}'].font = BOLD_FONT

    # Accounts Payable
    row += 1
    ap_row = row
    ws[f'A{row}'] = 'Accounts Payable'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        # AP = COGS * AP Days / 365
        ws[f'{col}{row}'] = f"='Income Statement'!{col}8*'Assumptions & Drivers'!{year_col}23/365"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Other Current Liabilities
    row += 1
    other_cl_row = row
    ws[f'A{row}'] = 'Other Current Liabilities'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"='Income Statement'!{col}7*'Assumptions & Drivers'!{year_col}25"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Total Current Liabilities
    row += 1
    total_cl_row = row
    ws[f'A{row}'] = 'Total Current Liabilities'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"=SUM({col}{ap_row}:{col}{other_cl_row})"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Revolver / Debt
    row += 1
    debt_row = row
    ws[f'A{row}'] = 'Revolver / Debt'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:  # 2021
            ws[f'{col}{row}'] = "='Assumptions & Drivers'!B30"
        else:
            # Link from Cash Flow
            ws[f'{col}{row}'] = f"='Cash Flow'!{col}30"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Total Liabilities
    row += 1
    total_liab_row = row
    ws[f'A{row}'] = 'Total Liabilities'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{total_cl_row}+{col}{debt_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Shareholders' Equity
    row += 1
    equity_row = row
    ws[f'A{row}'] = "Shareholders' Equity"
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:  # 2021
            ws[f'{col}{row}'] = "='Assumptions & Drivers'!B32"
        else:
            # Equity = Prior Equity + Net Income
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"={prev_col}{row}+'Income Statement'!{col}19"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Total Liabilities & Equity
    row += 1
    total_liab_eq_row = row
    ws[f'A{row}'] = 'Total Liabilities & Equity'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{total_liab_row}+{col}{equity_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Balance Check
    row += 2
    ws[f'A{row}'] = 'Balance Check'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{total_assets_row}-{col}{total_liab_eq_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        # Add conditional formatting would be nice but will skip for now

def build_cash_flow_statement(wb):
    """Build the Cash Flow Statement"""
    ws = wb["Cash Flow"]

    # Set column widths
    set_column_widths(ws, {
        'A': 30, 'B': 15, 'C': 12, 'D': 12, 'E': 12, 'F': 12,
        'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12
    })

    # Title
    ws['A1'] = 'CASH FLOW STATEMENT'
    ws['A1'].font = Font(size=14, bold=True)

    ws['A2'] = '($ in millions)'
    ws['A2'].font = Font(italic=True)

    # Years header
    row = 4
    ws[f'A{row}'] = 'Period'
    ws[f'A{row}'].font = BOLD_FONT

    col_start = 2  # Column B
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = year
        ws[f'{col}{row}'].font = HEADER_FONT
        ws[f'{col}{row}'].fill = HEADER_FILL
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # Operating Activities
    row += 1
    ws[f'A{row}'] = 'OPERATING ACTIVITIES'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # Net Income
    row += 1
    ws[f'A{row}'] = 'Net Income'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"='Income Statement'!{col}19"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # D&A
    row += 1
    ws[f'A{row}'] = 'D&A'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"='Income Statement'!{col}13"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Changes in Working Capital
    row += 1
    ws[f'A{row}'] = 'Changes in Working Capital:'
    ws[f'A{row}'].font = BOLD_FONT

    # Change in AR
    row += 1
    ws[f'A{row}'] = 'Change in AR'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = 0
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"=-('Balance Sheet'!{col}8-'Balance Sheet'!{prev_col}8)"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Change in Inventory
    row += 1
    ws[f'A{row}'] = 'Change in Inventory'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = 0
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"=-('Balance Sheet'!{col}9-'Balance Sheet'!{prev_col}9)"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Change in Other CA
    row += 1
    ws[f'A{row}'] = 'Change in Other Current Assets'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = 0
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"=-('Balance Sheet'!{col}10-'Balance Sheet'!{prev_col}10)"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Change in AP
    row += 1
    ws[f'A{row}'] = 'Change in AP'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = 0
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"='Balance Sheet'!{col}19-'Balance Sheet'!{prev_col}19"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Change in Other CL
    row += 1
    ws[f'A{row}'] = 'Change in Other Current Liab'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = 0
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"='Balance Sheet'!{col}20-'Balance Sheet'!{prev_col}20"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Cash from Operations
    row += 1
    cfo_row = row
    ws[f'A{row}'] = 'Cash from Operations'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"=SUM({col}6:{col}{row-1})"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Investing Activities
    row += 2
    ws[f'A{row}'] = 'INVESTING ACTIVITIES'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # CapEx
    row += 1
    capex_row = row
    ws[f'A{row}'] = 'CapEx'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        year_col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"=-('Income Statement'!{col}7*'Assumptions & Drivers'!{year_col}26)"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Cash from Investing
    row += 1
    cfi_row = row
    ws[f'A{row}'] = 'Cash from Investing'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{capex_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Financing Activities
    row += 2
    ws[f'A{row}'] = 'FINANCING ACTIVITIES'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # Net Borrowing (calculated to balance)
    row += 1
    net_borrow_row = row
    ws[f'A{row}'] = 'Net Borrowing / (Repayment)'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = 0
        else:
            prev_col = get_column_letter(col_start + i - 1)
            # Change in debt = ending debt - beginning debt
            ws[f'{col}{row}'] = f"='Balance Sheet'!{col}26-'Balance Sheet'!{prev_col}26"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Cash from Financing
    row += 1
    cff_row = row
    ws[f'A{row}'] = 'Cash from Financing'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{net_borrow_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Net Change in Cash
    row += 2
    ws[f'A{row}'] = 'Net Change in Cash'
    ws[f'A{row}'].font = BOLD_FONT
    net_change_row = row
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{cfo_row}+{col}{cfi_row}+{col}{cff_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Beginning Cash
    row += 1
    ws[f'A{row}'] = 'Beginning Cash'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = "='Assumptions & Drivers'!B33"
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"={prev_col}{row+1}"  # Link to prior ending cash
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Ending Cash (before revolver adjustment)
    row += 1
    ws[f'A{row}'] = 'Ending Cash (before revolver)'
    prelim_cash_row = row
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{row-1}+{col}{net_change_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Revolver logic
    row += 2
    ws[f'A{row}'] = 'REVOLVER LOGIC'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    # Beginning Debt
    row += 1
    beg_debt_row = row
    ws[f'A{row}'] = 'Beginning Debt'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        if i == 0:
            ws[f'{col}{row}'] = "='Assumptions & Drivers'!B30"
        else:
            prev_col = get_column_letter(col_start + i - 1)
            ws[f'{col}{row}'] = f"={prev_col}{row+1}"  # Link to prior ending debt
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Ending Debt (with revolver)
    row += 1
    end_debt_row = row
    ws[f'A{row}'] = 'Ending Debt'
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        # If cash would be negative, draw on revolver. If positive, pay down debt.
        ws[f'{col}{row}'] = f"=MAX(0,{col}{beg_debt_row}-{col}{prelim_cash_row})"
        ws[f'{col}{row}'].number_format = '#,##0.0'

    # Ending Cash (final)
    row += 1
    end_cash_row = row
    ws[f'A{row}'] = 'Ending Cash'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        # Cash = prelim cash + debt draw or - debt paydown
        ws[f'{col}{row}'] = f"={col}{prelim_cash_row}+{col}{beg_debt_row}-{col}{end_debt_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

    # Free Cash Flow
    row += 2
    ws[f'A{row}'] = 'KEY METRICS'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].fill = SECTION_FILL

    row += 1
    fcf_row = row
    ws[f'A{row}'] = 'Free Cash Flow'
    ws[f'A{row}'].font = BOLD_FONT
    for i, year in enumerate(ALL_YEARS):
        col = get_column_letter(col_start + i)
        ws[f'{col}{row}'] = f"={col}{cfo_row}+{col}{capex_row}"
        ws[f'{col}{row}'].number_format = '#,##0.0'
        ws[f'{col}{row}'].font = BOLD_FONT

def build_charts(wb):
    """Build the Charts sheet"""
    ws = wb["Charts"]

    # Set column widths
    set_column_widths(ws, {'A': 20, 'B': 15, 'C': 15, 'D': 15})

    # Title
    ws['A1'] = 'KEY CHARTS & VISUALIZATIONS'
    ws['A1'].font = Font(size=14, bold=True)

    # Data table for reference
    ws['A3'] = 'Year'
    ws['B3'] = 'Revenue'
    ws['C3'] = 'EBITDA'
    ws['D3'] = 'Free Cash Flow'

    for i, cell in enumerate(['A3', 'B3', 'C3', 'D3']):
        ws[cell].font = BOLD_FONT
        ws[cell].fill = HEADER_FILL

    # Populate data
    for i, year in enumerate(ALL_YEARS):
        row = 4 + i
        col_letter = get_column_letter(2 + i)

        ws[f'A{row}'] = year
        ws[f'B{row}'] = f"='Income Statement'!{col_letter}7"
        ws[f'C{row}'] = f"='Income Statement'!{col_letter}22"
        ws[f'D{row}'] = f"='Cash Flow'!{col_letter}36"

        ws[f'B{row}'].number_format = '#,##0.0'
        ws[f'C{row}'].number_format = '#,##0.0'
        ws[f'D{row}'].number_format = '#,##0.0'

    # Revenue Chart
    chart1 = LineChart()
    chart1.title = "Revenue Growth"
    chart1.style = 13
    chart1.y_axis.title = '$ millions'
    chart1.x_axis.title = 'Year'

    data1 = Reference(ws, min_col=2, min_row=3, max_row=3+len(ALL_YEARS))
    cats1 = Reference(ws, min_col=1, min_row=4, max_row=3+len(ALL_YEARS))
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)

    ws.add_chart(chart1, "F3")

    # EBITDA Chart
    chart2 = LineChart()
    chart2.title = "EBITDA Trend"
    chart2.style = 13
    chart2.y_axis.title = '$ millions'
    chart2.x_axis.title = 'Year'

    data2 = Reference(ws, min_col=3, min_row=3, max_row=3+len(ALL_YEARS))
    cats2 = Reference(ws, min_col=1, min_row=4, max_row=3+len(ALL_YEARS))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)

    ws.add_chart(chart2, "F18")

    # FCF Chart
    chart3 = LineChart()
    chart3.title = "Free Cash Flow"
    chart3.style = 13
    chart3.y_axis.title = '$ millions'
    chart3.x_axis.title = 'Year'

    data3 = Reference(ws, min_col=4, min_row=3, max_row=3+len(ALL_YEARS))
    cats3 = Reference(ws, min_col=1, min_row=4, max_row=3+len(ALL_YEARS))
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)

    ws.add_chart(chart3, "F33")

def main():
    """Main function to build the complete model"""
    print("Building 3-Statement Financial Model...")

    # Create workbook
    wb = create_workbook()
    print("✓ Workbook created")

    # Build each sheet
    print("Building Assumptions & Drivers sheet...")
    build_assumptions_sheet(wb)
    print("✓ Assumptions & Drivers complete")

    print("Building Income Statement...")
    build_income_statement(wb)
    print("✓ Income Statement complete")

    print("Building Balance Sheet...")
    build_balance_sheet(wb)
    print("✓ Balance Sheet complete")

    print("Building Cash Flow Statement...")
    build_cash_flow_statement(wb)
    print("✓ Cash Flow Statement complete")

    print("Building Charts...")
    build_charts(wb)
    print("✓ Charts complete")

    # Save the file
    filename = "3_Statement_Financial_Model.xlsx"
    wb.save(filename)
    print(f"\n✓ Model saved as {filename}")
    print("\nModel features:")
    print("  • Historical years: 2021-2024")
    print("  • Forecast years: 2025-2029")
    print("  • Revenue segments: Product & Service")
    print("  • Scenario analysis: Base/Upside/Downside")
    print("  • Full working capital drivers")
    print("  • PP&E roll-forward with CapEx")
    print("  • Integrated cash flow with revolver logic")
    print("  • Visual charts for Revenue, EBITDA, and FCF")
    print("\nColor coding:")
    print("  • Blue cells = Inputs")
    print("  • Black cells = Formulas")
    print("\nTo use: Change scenario in cell B4 of 'Assumptions & Drivers' sheet")

if __name__ == "__main__":
    main()
